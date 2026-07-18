"""Parse Insolvency Service company-insolvency data into JSON for the data hub page.

Reusable: re-run when a new monthly release arrives by swapping the CSV/XLSX
source files in `data/insolvency-statistics/`.

Outputs (written to data/insolvency-statistics/):
  - monthly_series.json         long-run monthly counts by procedure (E&W)
  - rate_series.json            12-month rolling rate by month (E&W)
  - sector_breakdown.json       12 months to latest sector month, top sections
  - uk_nations.json             latest-month counts & rates for E&W / SC / NI
  - release_metadata.json       publication date, latest month, next release

Cleaning rules (per brief):
  - [x] -> None (unavailable, do not chart)
  - [z] -> None (not applicable)
  - commas stripped before numeric parse
  - percentage changes not calculated if either value < 5
"""

from __future__ import annotations

import argparse
import csv
import json
from datetime import date
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "data" / "insolvency-statistics"

# ── SIC sections we surface in the sector breakdown ──────────────────────
# Section code -> short label for the chart, long label for the table
SECTION_LABELS = {
    "F": ("Construction", "Construction"),
    "G": ("Wholesale and retail", "Wholesale and retail trade; repair of motor vehicles and motorcycles"),
    "I": ("Accommodation and food", "Accommodation and food service activities"),
    "N": ("Admin and support", "Administrative and support service activities"),
    "M": ("Professional services", "Professional, scientific and technical activities"),
    "C": ("Manufacturing", "Manufacturing"),
    "H": ("Transport and storage", "Transport and storage"),
    "J": ("Information and communication", "Information and communication"),
    "K": ("Finance and insurance", "Financial and insurance activities"),
    "L": ("Real estate", "Real estate activities"),
    "S": ("Other services", "Other service activities"),
    "P": ("Education", "Education"),
    "Q": ("Health and social work", "Human health and social work activities"),
    "R": ("Arts and recreation", "Arts, entertainment and recreation"),
    "A": ("Agriculture", "Agriculture, forestry and fishing"),
    "B": ("Mining and quarrying", "Mining and quarrying"),
    "D": ("Electricity and gas", "Electricity, gas, steam and air conditioning supply"),
    "E": ("Water and waste", "Water supply; sewerage, waste management and remediation activities"),
    "O": ("Public administration", "Public administration and defence; compulsory social security"),
    "T": ("Households as employers", "Activities of households as employers"),
    "U": ("Extraterritorial", "Activities of extraterritorial organisations and bodies"),
}

MONTH_NAMES = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
               "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def clean_value(raw):
    """Apply the brief's cleaning rules. Returns float, int, or None."""
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return raw
    s = str(raw).strip()
    if s in ("", "[x]", "[z]", "[p]", "..", "-"):
        # [p] alone marks provisional but doesn't carry a value here
        return None
    s = s.replace(",", "").replace("[p]", "").strip()
    try:
        return float(s) if "." in s else int(s)
    except ValueError:
        return None


def pct_change(latest, prior, threshold=5):
    """Return rounded percent change, or None when either value < threshold."""
    if latest is None or prior is None:
        return None
    if abs(latest) < threshold or abs(prior) < threshold:
        return None
    if prior == 0:
        return None
    return round((latest - prior) / prior * 100)


def parse_long_run_csv(csv_path: Path) -> dict:
    """Read the long-run CSV. Returns a dict keyed by series with month lists."""
    months = []
    series = {
        "total_ew_sa": [],
        "compliq_ew_sa": [],
        "cvl_ew_sa": [],
        "admin_ew_sa": [],
        "cva_ew_nsa": [],
        "rec_ew_nsa": [],
        "tot_rate_ew": [],
        # NSA fallback for non-seasonally-adjusted comparison
        "total_ew_nsa": [],
        "compliq_ew_nsa": [],
        "cvl_ew_nsa": [],
        "admin_ew_nsa": [],
    }
    with csv_path.open(encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            months.append(row["period"])
            for key in series:
                series[key].append(clean_value(row.get(key)))
    return {"months": months, **series}


def parse_industry_sheet(xlsx_path: Path) -> dict:
    """Parse Table_1c — industry breakdown. Returns 12-month-to-latest section totals."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["Table_1c"]
    rows = list(ws.iter_rows(values_only=True))

    # Header row is the one starting with "Section"
    header_idx = next(i for i, r in enumerate(rows) if r and r[0] == "Section")
    header = rows[header_idx]

    # Locate monthly columns. Skip annual-total columns (numeric labels like 2016).
    month_cols = []
    for col_idx, label in enumerate(header):
        if not isinstance(label, str):
            continue
        for mi, mname in enumerate(MONTH_NAMES, start=1):
            if label.startswith(mname + " ") and len(label) == 8:
                try:
                    year = int(label[-4:])
                    month_cols.append((col_idx, year, mi, label))
                except ValueError:
                    pass
                break

    if not month_cols:
        raise RuntimeError("Could not locate monthly columns in Table_1c")

    # Determine the latest available month (industry data lags one month)
    # We want the 12 months ending at the latest non-empty month.
    # Scan rows looking for the Total row and find the last column with a value.
    total_row = next(r for r in rows if r and r[0] == "Total")
    available_months = []
    for col_idx, year, month, label in month_cols:
        if clean_value(total_row[col_idx]) is not None:
            available_months.append((col_idx, year, month, label))

    if len(available_months) < 12:
        raise RuntimeError(f"Need at least 12 months of industry data; have {len(available_months)}")

    window = available_months[-12:]   # the most recent 12 months with data
    window_cols = [c for c, *_ in window]

    section_totals: dict[str, int] = {}
    for r in rows:
        if not r or not isinstance(r[0], str):
            continue
        section = r[0].strip()
        division = r[1]
        group = r[2]
        # Section-level rows have empty Division and Group
        if section == "Total" or section not in SECTION_LABELS:
            continue
        if division is not None or group is not None:
            continue
        total = 0
        ok = False
        for c in window_cols:
            v = clean_value(r[c])
            if v is not None:
                total += int(v)
                ok = True
        if ok:
            section_totals[section] = total

    grand_total_known_sector = sum(section_totals.values())
    breakdown = []
    for code, total in sorted(section_totals.items(), key=lambda kv: -kv[1]):
        short, long_label = SECTION_LABELS[code]
        share_pct = round(total / grand_total_known_sector * 100) if grand_total_known_sector else 0
        breakdown.append({
            "code": code,
            "label_short": short,
            "label_full": long_label,
            "count": total,
            "share_pct": share_pct,
        })

    latest_label = window[-1][3]
    earliest_label = window[0][3]

    return {
        "window_start": earliest_label,
        "window_end": latest_label,
        "total_known_sector": grand_total_known_sector,
        "sections": breakdown,
    }


def parse_nations(xlsx_path: Path) -> dict:
    """Read latest-month figures and rates for Scotland and Northern Ireland.

    Tables 4a, 5, 6, 7 share the same column layout:
      col 0: Period (year string or "Mmm YYYY")
      col 1: Total Company Insolvencies   — counts in 4a/6, rates in 5/7
      cols 2..6: by procedure (counts or rates)
    The bottom of each sheet contains "Percentage change" summary rows we ignore.
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    def looks_like_month(period_str: str) -> bool:
        """True for 'Apr 2026' style labels — filters out 'vs Apr 2025' etc."""
        parts = period_str.strip().split()
        return len(parts) == 2 and parts[0] in MONTH_NAMES and parts[1].isdigit()

    def latest_monthly_row(sheet_name: str, value_caster):
        """Return (period_label, value) for the latest month row in the sheet."""
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        latest_period, latest_value = None, None
        for r in rows:
            if not r or not isinstance(r[0], str):
                continue
            period = r[0].strip()
            if not looks_like_month(period):
                continue
            v = clean_value(r[1])  # col 1 = total (count or rate, by sheet)
            if v is not None:
                latest_period = period
                latest_value = value_caster(v)
        return latest_period, latest_value

    sc_period, sc_total = latest_monthly_row("Table_4a", int)
    ni_period, ni_total = latest_monthly_row("Table_6", int)
    _, sc_rate = latest_monthly_row("Table_5", float)
    _, ni_rate = latest_monthly_row("Table_7", float)

    return {
        "scotland": {"period": sc_period, "total": sc_total, "rate_per_10k": sc_rate},
        "northern_ireland": {"period": ni_period, "total": ni_total, "rate_per_10k": ni_rate},
    }


def build_release_metadata(months: list[str]) -> dict:
    """Release metadata. Hard-coded values match June 2026 release."""
    return {
        "latest_month": months[-1] if months else None,
        "latest_month_label": "June 2026",
        "publication_date": "17 July 2026",
        "next_release_date": "18 August 2026",
        "source_label": "Insolvency Service / Companies House",
        "status": "Accredited official statistics",
        "data_window_start": months[0] if months else None,
        "page_generated": date.today().isoformat(),
    }


def build_latest_figures_table(long_run: dict) -> dict:
    """April / March / April-prior breakdown for the latest-figures table."""
    months = long_run["months"]
    if len(months) < 13:
        raise RuntimeError("Need at least 13 months of data for YoY comparison")
    latest_idx = len(months) - 1
    prior_month_idx = latest_idx - 1
    prior_year_idx = latest_idx - 12

    rows = []
    procedures = [
        ("Total company insolvencies", "total_ew_sa"),
        ("Compulsory liquidations", "compliq_ew_sa"),
        ("Creditors' voluntary liquidations", "cvl_ew_sa"),
        ("Administrations", "admin_ew_sa"),
        ("CVAs", "cva_ew_nsa"),
        ("Receivership appointments", "rec_ew_nsa"),
    ]
    for label, series_key in procedures:
        latest = long_run[series_key][latest_idx]
        prev_m = long_run[series_key][prior_month_idx]
        prev_y = long_run[series_key][prior_year_idx]
        rows.append({
            "procedure": label,
            "latest": latest,
            "prior_month": prev_m,
            "prior_year": prev_y,
            "month_change_pct": pct_change(latest, prev_m),
            "year_change_pct": pct_change(latest, prev_y),
        })
    return {
        "latest_month_label": months[latest_idx],
        "prior_month_label": months[prior_month_idx],
        "prior_year_label": months[prior_year_idx],
        "rows": rows,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Parse Insolvency Service data into JSON")
    parser.add_argument("--csv", default=str(DATA_DIR / "source_long_run_2026-06.csv"))
    parser.add_argument("--xlsx", default=str(DATA_DIR / "source_data_tables_2026-06.xlsx"))
    parser.add_argument("--out-dir", default=str(DATA_DIR))
    args = parser.parse_args()

    csv_path = Path(args.csv)
    xlsx_path = Path(args.xlsx)
    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    print(f"Reading CSV: {csv_path}")
    long_run = parse_long_run_csv(csv_path)

    print(f"Reading XLSX: {xlsx_path}")
    sectors = parse_industry_sheet(xlsx_path)
    nations = parse_nations(xlsx_path)

    metadata = build_release_metadata(long_run["months"])
    latest_table = build_latest_figures_table(long_run)

    # ── monthly_series.json: full long-run for charts ─────────────────────
    monthly_series = {
        "months": long_run["months"],
        "series": {
            "cvls":          long_run["cvl_ew_sa"],
            "compulsory":    long_run["compliq_ew_sa"],
            "administrations": long_run["admin_ew_sa"],
            "cvas":          long_run["cva_ew_nsa"],
            "receiverships": long_run["rec_ew_nsa"],
            "total":         long_run["total_ew_sa"],
        },
        "note": "England and Wales. CVLs, compulsory liquidations and administrations are seasonally adjusted. CVAs and receiverships are not seasonally adjusted due to low volumes.",
    }
    (out_dir / "monthly_series.json").write_text(json.dumps(monthly_series, indent=2))

    # ── rate_series.json ──────────────────────────────────────────────────
    rate_series = {
        "months": long_run["months"],
        "rate_per_10k": long_run["tot_rate_ew"],
        "note": "12-month rolling rate, England and Wales, per 10,000 active companies.",
    }
    (out_dir / "rate_series.json").write_text(json.dumps(rate_series, indent=2))

    # ── sector_breakdown.json ─────────────────────────────────────────────
    (out_dir / "sector_breakdown.json").write_text(json.dumps(sectors, indent=2))

    # ── uk_nations.json ───────────────────────────────────────────────────
    nations_full = {
        "england_and_wales": {
            "period": long_run["months"][-1],
            "total": long_run["total_ew_sa"][-1],
            "rate_per_10k": long_run["tot_rate_ew"][-1],
        },
        **nations,
    }
    (out_dir / "uk_nations.json").write_text(json.dumps(nations_full, indent=2))

    # ── release_metadata.json + latest figures table ──────────────────────
    metadata["latest_figures_table"] = latest_table
    (out_dir / "release_metadata.json").write_text(json.dumps(metadata, indent=2))

    print(f"\nWrote outputs to {out_dir}")
    print(f"  - Latest month: {metadata['latest_month']}")
    print(f"  - E&W total: {long_run['total_ew_sa'][-1]}")
    print(f"  - Rate: {long_run['tot_rate_ew'][-1]} per 10,000")
    print(f"  - Sector window: {sectors['window_start']} to {sectors['window_end']}")
    print(f"  - Top sector: {sectors['sections'][0]['label_short']} = {sectors['sections'][0]['count']:,} ({sectors['sections'][0]['share_pct']}%)")
    print(f"  - Scotland latest: {nations['scotland']}")
    print(f"  - NI latest: {nations['northern_ireland']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
